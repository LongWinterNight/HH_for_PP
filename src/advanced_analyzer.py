"""
Модуль для расширенной аналитики вакансий.

Реализует:
- Подсчёт навыков по расширенным категориям (LLM, RAG, Vector DB и др.)
- Детализацию связей "вакансия-навык"
- Аналитику по группировкам как в ручном анализе
- Формирование детальных отчётов в Excel
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Set
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference

from src.config import config_loader, settings
from src.utils import get_logger, ensure_dir, safe_divide

logger = get_logger(__name__)


class AdvancedAnalytics:
    """
    Расширенный аналитический модуль для детального анализа вакансий.

    Реализует:
    - Подсчёт навыков по расширенным категориям из config.yaml
    - Связи "вакансия-навык" с указанием источника
    - Группировки по технологиям, hard skills, soft skills
    - Детальные Excel-отчёты с разбивкой по категориям

    Пример использования:
        >>> analytics = AdvancedAnalytics(df)
        >>> analytics.compute_category_statistics()
        >>> analytics.generate_detailed_report()
    """

    def __init__(self, df: pd.DataFrame) -> None:
        """
        Инициализация аналитического модуля.

        Args:
            df: DataFrame с обработанными вакансиями
        """
        self.df = df
        self.reports_dir = settings.reports_dir
        ensure_dir(self.reports_dir)

        # Загружаем расширенные категории
        self.advanced_categories = config_loader.get_advanced_category_skills()

        # Кэш для статистики
        self._category_stats: Optional[Dict[str, Any]] = None
        self._vacancy_skill_map: Optional[pd.DataFrame] = None

        logger.info(f"AdvancedAnalytics инициализирован. Вакансий: {len(df)}")

    def _count_skills_in_column(
        self,
        column: pd.Series,
        skill_set: Set[str]
    ) -> Dict[str, int]:
        """
        Подсчёт упоминаний навыков из заданного множества.

        Args:
            column: Серия с навыками (строки через запятую)
            skill_set: Множество навыков для поиска

        Returns:
            Словарь {навык: частота}
        """
        skill_counts: Dict[str, int] = defaultdict(int)

        for skills_str in column.dropna():
            if isinstance(skills_str, str):
                skills = [s.strip().lower() for s in skills_str.split(",")]
                for skill in skills:
                    if skill in skill_set:
                        skill_counts[skill] += 1

        return dict(skill_counts)

    def _normalize_skill_name(self, skill: str) -> str:
        """Нормализация названия навыка для отображения."""
        # Capitalize первое слово
        parts = skill.split()
        return " ".join([parts[0].capitalize()] + parts[1:]) if parts else skill

    def compute_category_statistics(self) -> Dict[str, Dict[str, int]]:
        """
        Подсчёт статистики по расширенным категориям.

        Returns:
            Словарь {category: {skill: count}}
        """
        if self._category_stats is not None:
            return self._category_stats

        if self.df.empty:
            return {}

        category_stats = {}

        # Объединяем все колонки навыков в одну для анализа
        all_skills_text = (
            self.df["hard_skills"].fillna("") + ", " +
            self.df["soft_skills"].fillna("") + ", " +
            self.df["tools"].fillna("")
        )

        # Создаём фейковую серию для совместимости с _count_skills_in_column
        fake_series = pd.Series([all_skills_text.iloc[i] for i in range(len(self.df))])

        # Подсчитываем по каждой категории
        for category_name, skills_list in self.advanced_categories.items():
            # Нормализуем названия навыков
            normalized_skills = {s.lower().strip() for s in skills_list}

            # Считаем упоминания
            counts = self._count_skills_in_column(
                pd.Series([", ".join([
                    skill for skill in normalized_skills
                    if skill in all_skills_text.iloc[i].lower()
                ]) for i in range(len(self.df))]),
                normalized_skills
            )

            # Более точный подсчёт через прямой поиск
            counts = {}
            for skill in skills_list:
                skill_lower = skill.lower()
                count = 0
                for _, row in self.df.iterrows():
                    # Проверяем все колонки навыков
                    for col in ["hard_skills", "soft_skills", "tools"]:
                        skills_in_col = str(row.get(col, "")).lower().split(",")
                        skills_in_col = [s.strip() for s in skills_in_col]
                        if skill_lower in skills_in_col:
                            count += 1
                            break
                if count > 0:
                    counts[skill] = count

            # Сортируем по убыванию
            category_stats[category_name] = dict(sorted(
                counts.items(),
                key=lambda x: x[1],
                reverse=True
            ))

        self._category_stats = category_stats
        return category_stats

    def compute_vacancy_skill_mapping(self) -> pd.DataFrame:
        """
        Создание детальной карты связей "вакансия-навык".

        Returns:
            DataFrame с колонками:
            - vacancy_id, vacancy_name, area, employer_name
            - skill_name, skill_category, advanced_category
            - published_at, vacancy_url
        """
        if self._vacancy_skill_map is not None:
            return self._vacancy_skill_map

        if self.df.empty:
            return pd.DataFrame()

        rows = []

        # Создаём обратный индекс для расширенных категорий
        skill_to_advanced_category = {}
        for category_name, skills_list in self.advanced_categories.items():
            for skill in skills_list:
                skill_to_advanced_category[skill.lower().strip()] = category_name

        # Проходим по каждой вакансии
        for _, row in self.df.iterrows():
            vacancy_info = {
                "vacancy_id": row.get("vacancy_id"),
                "vacancy_name": row.get("vacancy_name", ""),
                "area": row.get("area", ""),
                "employer_name": row.get("employer_name", ""),
                "published_at": row.get("published_at"),
                "vacancy_url": row.get("vacancy_url", ""),
                "experience": row.get("experience", ""),
                "salary_from": row.get("salary_from"),
            }

            # Проходим по всем категориям навыков
            for category, col_name in [
                ("hard_skills", "hard_skills"),
                ("soft_skills", "soft_skills"),
                ("tools", "tools")
            ]:
                skills_str = row.get(col_name, "")
                if not skills_str or not isinstance(skills_str, str):
                    continue

                skills = [s.strip() for s in skills_str.split(",") if s.strip()]

                for skill in skills:
                    skill_row = vacancy_info.copy()
                    skill_row["skill_name"] = skill
                    skill_row["skill_category"] = category
                    skill_row["advanced_category"] = skill_to_advanced_category.get(
                        skill.lower(), ""
                    )
                    rows.append(skill_row)

        self._vacancy_skill_map = pd.DataFrame(rows)
        logger.info(f"Создана карта связей: {len(self._vacancy_skill_map)} записей")

        return self._vacancy_skill_map

    def compute_technology_summary(self) -> Dict[str, Dict[str, Any]]:
        """
        Сводка по технологиям с группировкой как в ручном анализе.

        Returns:
            Словарь с группировками:
            - llm_tools: LLM и инструменты генерации
            - vector_db: Векторные базы
            - rag: RAG технологии
            - ml_libs: ML библиотеки
            - databases: Базы данных
            - cloud: Облачные платформы
            - и другие
        """
        if self.df.empty:
            return {}

        # Группировки технологий (как в вашем анализе)
        tech_groups = {
            "LLM и Генеративный ИИ / API": {
                "skills": ["llm", "gpt", "gpt-4", "generative ai", "genai", "openai api",
                          "anthropic", "claude", "llama", "stable diffusion", "midjourney"],
                "count": 0
            },
            "Python": {
                "skills": ["python"],
                "count": 0
            },
            "REST API / Webhooks / API-интеграции": {
                "skills": ["rest api", "rest", "api", "webhooks", "api integration", "http"],
                "count": 0
            },
            "LLM-фреймворки": {
                "skills": ["langchain", "langgraph", "llama index", "transformers", "hugging face"],
                "count": 0
            },
            "Векторные БД / Vector Search": {
                "skills": ["vector database", "vector search", "pinecone", "weaviate",
                          "milvus", "qdrant", "chromadb", "faiss", "embeddings"],
                "count": 0
            },
            "RAG (вкл. CRAG, GraphRAG)": {
                "skills": ["rag", "retrieval augmented generation", "graphrag", "crag"],
                "count": 0
            },
            "AI-агенты / Workflows / AI Automation": {
                "skills": ["ai agents", "ai automation", "workflows", "autonomous agents",
                          "multi-agent", "agent pipeline"],
                "count": 0
            },
            "Python Веб-фреймворки": {
                "skills": ["fastapi", "flask", "django", "aiohttp"],
                "count": 0
            },
            "Базы данных (SQL / Postgres / SQLite)": {
                "skills": ["sql", "postgresql", "mysql", "sqlite", "clickhouse"],
                "count": 0
            },
            "Docker / Контейнеризация": {
                "skills": ["docker", "kubernetes", "k8s", "docker-compose"],
                "count": 0
            },
            "Облачные платформы": {
                "skills": ["aws", "azure", "gcp", "google cloud", "s3", "yandex cloud"],
                "count": 0
            },
            "ML Библиотеки": {
                "skills": ["pandas", "numpy", "scikit-learn", "pytorch", "tensorflow",
                          "keras", "xgboost", "lightgbm", "catboost"],
                "count": 0
            },
            "Краудсорсинг / Разметка": {
                "skills": ["toloka", "яндекс толока", "crowdsourcing", "data annotation",
                          "разметка данных", "label studio"],
                "count": 0
            },
            "CI/CD": {
                "skills": ["ci/cd", "jenkins", "gitlab ci", "github actions"],
                "count": 0
            },
            "Airflow": {
                "skills": ["airflow", "apache airflow"],
                "count": 0
            },
            "Kubernetes": {
                "skills": ["kubernetes", "k8s", "helm"],
                "count": 0
            },
            "Linux": {
                "skills": ["linux", "bash"],
                "count": 0
            },
            "Графовые БД": {
                "skills": ["graph database", "neo4j", "arangodb", "knowledge graph"],
                "count": 0
            },
        }

        # Подсчитываем упоминания по каждой группе
        for group_name, group_data in tech_groups.items():
            group_skills = set(group_data["skills"])
            count = 0

            for _, row in self.df.iterrows():
                # Проверяем все колонки навыков
                found_in_vacancy = False
                for col in ["hard_skills", "soft_skills", "tools"]:
                    skills_str = row.get(col, "")
                    if not skills_str or not isinstance(skills_str, str):
                        continue
                    skills = [s.lower().strip() for s in skills_str.split(",")]
                    if any(skill in group_skills for skill in skills):
                        found_in_vacancy = True
                        break

                if found_in_vacancy:
                    count += 1

            group_data["count"] = count

        # Сортируем по количеству упоминаний
        sorted_groups = dict(sorted(
            tech_groups.items(),
            key=lambda x: x[1]["count"],
            reverse=True
        ))

        return sorted_groups

    def compute_hard_skills_summary(self) -> Dict[str, Dict[str, Any]]:
        """
        Сводка по Hard Skills с группировкой.

        Returns:
            Словарь с группировками hard skills
        """
        if self.df.empty:
            return {}

        hard_skill_groups = {
            "Работа с метриками / Оценка (Eval) / Бенчмарки / KPI": {
                "skills": ["metrics", "kpi", "evaluation", "eval", "benchmarks",
                          "бенчмарки", "оценка качества"],
                "count": 0
            },
            "Запуск от PoC до Production / MVP / Масштабирование": {
                "skills": ["mvp", "poc", "proof of concept", "production",
                          "масштабирование", "scaling", "deployment"],
                "count": 0
            },
            "Product / System / Business Analysis / Сбор требований": {
                "skills": ["product analysis", "system analysis", "business analysis",
                          "requirements gathering", "сбор требований", "системный дизайн"],
                "count": 0
            },
            "Prompt Engineering / Системные промпты": {
                "skills": ["prompt engineering", "prompt design", "системные промпты"],
                "count": 0
            },
            "Архитектура, Системный дизайн ИИ, Интеграция": {
                "skills": ["architecture", "system design", "архитектура",
                          "системная интеграция", "ai architecture"],
                "count": 0
            },
            "Построение Data / ML / ETL / Agent пайплайнов": {
                "skills": ["etl", "data pipeline", "ml pipeline", "agent pipeline",
                          "data engineering", "ml ops"],
                "count": 0
            },
            "Написание тех. документации / Спецификаций": {
                "skills": ["technical documentation", "техническая документация",
                          "specifications", "спецификации", "technical writing"],
                "count": 0
            },
            "QA / Оптимизация / Тестирование / Разметка данных": {
                "skills": ["qa", "quality assurance", "тестирование", "testing",
                          "data annotation", "разметка данных", "optimization"],
                "count": 0
            },
            "RAG-инженерия (Чанкинг, эмбеддинги, Knowledge Graphs)": {
                "skills": ["rag", "chunking", "чанкинг", "embeddings", "эмбеддинги",
                          "knowledge graph", "граф знаний"],
                "count": 0
            },
            "Математическая статистика / Линал / Теорвер": {
                "skills": ["математическая статистика", "теория вероятностей",
                          "линейная алгебра", "statistics", "linear algebra"],
                "count": 0
            },
            "Fine-tuning / RL / Подготовка датасетов": {
                "skills": ["fine-tuning", "reinforcement learning", "rl",
                          "обучение с подкреплением", "dataset preparation"],
                "count": 0
            },
        }

        # Подсчитываем упоминания
        for group_name, group_data in hard_skill_groups.items():
            group_skills = set(group_data["skills"])
            count = 0

            for _, row in self.df.iterrows():
                found_in_vacancy = False
                for col in ["hard_skills", "soft_skills", "tools"]:
                    skills_str = row.get(col, "")
                    if not skills_str or not isinstance(skills_str, str):
                        continue
                    skills = [s.lower().strip() for s in skills_str.split(",")]
                    if any(skill in group_skills for skill in skills):
                        found_in_vacancy = True
                        break

                if found_in_vacancy:
                    count += 1

            group_data["count"] = count

        # Сортируем
        return dict(sorted(
            hard_skill_groups.items(),
            key=lambda x: x[1]["count"],
            reverse=True
        ))

    def compute_soft_skills_summary(self) -> Dict[str, Dict[str, Any]]:
        """
        Сводка по Soft Skills с группировкой.

        Returns:
            Словарь с группировками soft skills
        """
        if self.df.empty:
            return {}

        soft_skill_groups = {
            "Нацеленность на результат / Прагматичность / Ownership": {
                "skills": ["result oriented", "нацеленность на результат", "pragmatic",
                          "прагматичность", "ownership", "ответственность"],
                "count": 0
            },
            "Работа в команде / Коллаборация": {
                "skills": ["teamwork", "командная работа", "collaboration",
                          "коллаборация", "cross-functional"],
                "count": 0
            },
            "Коммуникабельность / Стейкхолдер-менеджмент": {
                "skills": ["communication", "коммуникабельность", "stakeholder management",
                          "стейкхолдер-менеджмент"],
                "count": 0
            },
            "Стремление к развитию / Эксперименты / Креативность": {
                "skills": ["creativity", "креативность", "experiments", "эксперименты",
                          "self-development", "саморазвитие", "continuous learning"],
                "count": 0
            },
            "Внимательность к деталям": {
                "skills": ["attention to detail", "внимательность", "detail-oriented"],
                "count": 0
            },
            "Startup Mindset / Адаптивность / Скорость": {
                "skills": ["startup mindset", "agile", "scrum", "fast-paced",
                          "адаптивность", "adaptability", "гибкость"],
                "count": 0
            },
            "Обучение пользователей / Менторство / Гайдлайны": {
                "skills": ["user training", "обучение пользователей", "mentoring",
                          "менторство", "guidelines", "гайдлайны"],
                "count": 0
            },
            "Кросс-функциональная координация": {
                "skills": ["cross-functional coordination", "кросс-функциональная координация",
                          "coordination", "координация"],
                "count": 0
            },
        }

        # Подсчитываем упоминания
        for group_name, group_data in soft_skill_groups.items():
            group_skills = set(group_data["skills"])
            count = 0

            for _, row in self.df.iterrows():
                found_in_vacancy = False
                for col in ["hard_skills", "soft_skills", "tools"]:
                    skills_str = row.get(col, "")
                    if not skills_str or not isinstance(skills_str, str):
                        continue
                    skills = [s.lower().strip() for s in skills_str.split(",")]
                    if any(skill in group_skills for skill in skills):
                        found_in_vacancy = True
                        break

                if found_in_vacancy:
                    count += 1

            group_data["count"] = count

        # Сортируем
        return dict(sorted(
            soft_skill_groups.items(),
            key=lambda x: x[1]["count"],
            reverse=True
        ))

    def generate_detailed_excel_report(
        self,
        filename: Optional[str] = None
    ) -> Path:
        """
        Генерация детального Excel-отчёта с расширенной аналитикой.

        Листы:
        1. Summary — сводка по категориям
        2. Technology Groups — технологии с группировкой
        3. Hard Skills Groups — hard skills с группировкой
        4. Soft Skills Groups — soft skills с группировкой
        5. Vacancy-Skill Map — детальная карта связей
        6. Advanced Categories — расширенные категории

        Args:
            filename: Имя файла (опционально)

        Returns:
            Путь к сохранённому файлу
        """
        if self.df.empty:
            raise ValueError("DataFrame пуст")

        # Генерируем имя файла
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"hh_advanced_analytics_{timestamp}.xlsx"

        filepath = self.reports_dir / filename
        logger.info(f"Генерация детального отчёта: {filepath}")

        # Создаём workbook
        wb = Workbook()
        wb.remove(wb.active)

        # Вычисляем статистику
        tech_summary = self.compute_technology_summary()
        hard_skills_summary = self.compute_hard_skills_summary()
        soft_skills_summary = self.compute_soft_skills_summary()
        category_stats = self.compute_category_statistics()
        vacancy_skill_map = self.compute_vacancy_skill_mapping()

        # Создаём листы
        self._create_advanced_summary_sheet(
            wb, tech_summary, hard_skills_summary, soft_skills_summary
        )
        self._create_technology_groups_sheet(wb, tech_summary)
        self._create_hard_skills_groups_sheet(wb, hard_skills_summary)
        self._create_soft_skills_groups_sheet(wb, soft_skills_summary)
        self._create_vacancy_skill_map_sheet(wb, vacancy_skill_map)
        self._create_advanced_categories_sheet(wb, category_stats)

        # Сохраняем
        wb.save(filepath)
        logger.info(f"Отчёт сохранён: {filepath}")

        return filepath

    def _create_advanced_summary_sheet(
        self,
        wb: Workbook,
        tech_summary: Dict,
        hard_skills_summary: Dict,
        soft_skills_summary: Dict
    ) -> None:
        """Создание сводного листа."""
        ws = wb.create_sheet("Summary")

        # Заголовок
        ws["A1"] = "📊 АНАЛИТИЧЕСКАЯ СВОДКА"
        ws["A1"].font = Font(bold=True, size=18)

        ws["A3"] = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws["A3"].font = Font(italic=True)

        ws["A4"] = f"Всего вакансий: {len(self.df)}"
        ws["A4"].font = Font(bold=True)

        # Топ технологий
        row = 6
        ws[f"A{row}"] = "🛠 Топ Технологий и Инструментов"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        row += 1

        ws[f"A{row}"] = "Категория"
        ws[f"B{row}"] = "Упоминаний"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].font = Font(bold=True)
        row += 1

        for group_name, data in list(tech_summary.items())[:15]:
            ws[f"A{row}"] = group_name
            ws[f"B{row}"] = data["count"]
            row += 1

        # Топ Hard Skills
        row += 2
        ws[f"A{row}"] = "🧠 Топ Hard Skills"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        row += 1

        ws[f"A{row}"] = "Категория"
        ws[f"B{row}"] = "Упоминаний"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].font = Font(bold=True)
        row += 1

        for group_name, data in list(hard_skills_summary.items())[:10]:
            ws[f"A{row}"] = group_name
            ws[f"B{row}"] = data["count"]
            row += 1

        # Топ Soft Skills
        row += 2
        ws[f"A{row}"] = "🤝 Топ Soft Skills"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        row += 1

        ws[f"A{row}"] = "Категория"
        ws[f"B{row}"] = "Упоминаний"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].font = Font(bold=True)
        row += 1

        for group_name, data in list(soft_skills_summary.items())[:8]:
            ws[f"A{row}"] = group_name
            ws[f"B{row}"] = data["count"]
            row += 1

    def _create_technology_groups_sheet(
        self,
        wb: Workbook,
        tech_summary: Dict
    ) -> None:
        """Лист с группировкой технологий."""
        ws = wb.create_sheet("Technology Groups")

        ws["A1"] = "🛠 Топ Технологий и Инструментов"
        ws["A1"].font = Font(bold=True, size=16)

        ws["A3"] = "Категория"
        ws["B3"] = "Упоминаний"
        ws["C3"] = "% от вакансий"
        ws["A3"].font = Font(bold=True)
        ws["B3"].font = Font(bold=True)
        ws["C3"].font = Font(bold=True)

        total = len(self.df)

        for i, (group_name, data) in enumerate(tech_summary.items(), start=4):
            ws[f"A{i}"] = group_name
            ws[f"B{i}"] = data["count"]
            ws[f"C{i}"] = f"{safe_divide(data['count'] * 100, total):.1f}%"

        # Таблица
        if tech_summary:
            tab = Table(
                displayName="TechGroupsTable",
                ref=f"A3:C{3 + len(tech_summary)}"
            )
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showRowStripes=True
            )
            ws.add_table(tab)

    def _create_hard_skills_groups_sheet(
        self,
        wb: Workbook,
        hard_skills_summary: Dict
    ) -> None:
        """Лист с группировкой Hard Skills."""
        ws = wb.create_sheet("Hard Skills Groups")

        ws["A1"] = "🧠 Топ Hard Skills"
        ws["A1"].font = Font(bold=True, size=16)

        ws["A3"] = "Категория"
        ws["B3"] = "Упоминаний"
        ws["C3"] = "% от вакансий"
        ws["A3"].font = Font(bold=True)
        ws["B3"].font = Font(bold=True)
        ws["C3"].font = Font(bold=True)

        total = len(self.df)

        for i, (group_name, data) in enumerate(hard_skills_summary.items(), start=4):
            ws[f"A{i}"] = group_name
            ws[f"B{i}"] = data["count"]
            ws[f"C{i}"] = f"{safe_divide(data['count'] * 100, total):.1f}%"

        if hard_skills_summary:
            tab = Table(
                displayName="HardSkillsGroupsTable",
                ref=f"A3:C{3 + len(hard_skills_summary)}"
            )
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showRowStripes=True
            )
            ws.add_table(tab)

    def _create_soft_skills_groups_sheet(
        self,
        wb: Workbook,
        soft_skills_summary: Dict
    ) -> None:
        """Лист с группировкой Soft Skills."""
        ws = wb.create_sheet("Soft Skills Groups")

        ws["A1"] = "🤝 Топ Soft Skills"
        ws["A1"].font = Font(bold=True, size=16)

        ws["A3"] = "Категория"
        ws["B3"] = "Упоминаний"
        ws["C3"] = "% от вакансий"
        ws["A3"].font = Font(bold=True)
        ws["B3"].font = Font(bold=True)
        ws["C3"].font = Font(bold=True)

        total = len(self.df)

        for i, (group_name, data) in enumerate(soft_skills_summary.items(), start=4):
            ws[f"A{i}"] = group_name
            ws[f"B{i}"] = data["count"]
            ws[f"C{i}"] = f"{safe_divide(data['count'] * 100, total):.1f}%"

        if soft_skills_summary:
            tab = Table(
                displayName="SoftSkillsGroupsTable",
                ref=f"A3:C{3 + len(soft_skills_summary)}"
            )
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showRowStripes=True
            )
            ws.add_table(tab)

    def _create_vacancy_skill_map_sheet(
        self,
        wb: Workbook,
        vacancy_skill_map: pd.DataFrame
    ) -> None:
        """Лист с детальной картой связей вакансия-навык."""
        ws = wb.create_sheet("Vacancy-Skill Map")

        ws["A1"] = "📋 Детализация: Вакансия — Навык"
        ws["A1"].font = Font(bold=True, size=16)

        if vacancy_skill_map.empty:
            ws["A3"] = "Нет данных"
            return

        # Заголовки
        headers = [
            "Vacancy ID",
            "Вакансия",
            "Компания",
            "Регион",
            "Опыт",
            "Зарплата (от)",
            "Навык",
            "Категория",
            "Расширенная категория"
        ]

        for col, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col, value=header)
            ws.cell(row=3, column=col).font = Font(bold=True)

        # Данные (ограничим первыми 1000 записями для производительности)
        data_to_write = vacancy_skill_map.head(1000)

        for i, (_, row) in enumerate(data_to_write.iterrows(), start=4):
            ws.cell(row=i, column=1, value=row.get("vacancy_id"))
            ws.cell(row=i, column=2, value=row.get("vacancy_name", "")[:50])
            ws.cell(row=i, column=3, value=row.get("employer_name", "")[:30])
            ws.cell(row=i, column=4, value=row.get("area", ""))
            ws.cell(row=i, column=5, value=row.get("experience", ""))
            ws.cell(row=i, column=6, value=row.get("salary_from"))
            ws.cell(row=i, column=7, value=row.get("skill_name"))
            ws.cell(row=i, column=8, value=row.get("skill_category", ""))
            ws.cell(row=i, column=9, value=row.get("advanced_category", ""))

        # Таблица
        tab = Table(
            displayName="VacancySkillMapTable",
            ref=f"A3:I{3 + len(data_to_write)}"
        )
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True
        )
        ws.add_table(tab)

        ws["A2"] = f"Всего связей: {len(vacancy_skill_map)} (показано: {min(len(vacancy_skill_map), 1000)})"

    def _create_advanced_categories_sheet(
        self,
        wb: Workbook,
        category_stats: Dict
    ) -> None:
        """Лист с расширенными категориями."""
        ws = wb.create_sheet("Advanced Categories")

        ws["A1"] = "📊 Расширенные категории навыков"
        ws["A1"].font = Font(bold=True, size=16)

        row = 3
        for category_name, skills_counts in category_stats.items():
            if not skills_counts:
                continue

            # Название категории
            ws[f"A{row}"] = category_name.replace("_", " ").title()
            ws[f"A{row}"].font = Font(bold=True, size=12)
            row += 1

            # Заголовки
            ws[f"A{row}"] = "Навык"
            ws[f"B{row}"] = "Упоминаний"
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"].font = Font(bold=True)
            row += 1

            # Навыки
            for skill, count in list(skills_counts.items())[:20]:
                ws[f"A{row}"] = skill
                ws[f"B{row}"] = count
                row += 1

            row += 1  # Пустая строка между категориями

    def print_advanced_summary(self) -> None:
        """Вывод расширенной сводки в консоль."""
        if self.df.empty:
            print("❌ Нет данных для анализа")
            return

        print("\n" + "=" * 70)
        print("📊 АНАЛИТИЧЕСКАЯ СВОДКА (ПО ВАКАНСИЯМ)")
        print("=" * 70)

        tech_summary = self.compute_technology_summary()
        hard_skills_summary = self.compute_hard_skills_summary()
        soft_skills_summary = self.compute_soft_skills_summary()

        # Топ технологий
        print("\n🛠 Топ Технологий и Инструментов:")
        for group_name, data in list(tech_summary.items())[:15]:
            print(f"   {group_name} — ({data['count']})")

        # Топ Hard Skills
        print("\n🧠 Топ Hard Skills:")
        for group_name, data in list(hard_skills_summary.items())[:12]:
            print(f"   {group_name} — ({data['count']})")

        # Топ Soft Skills
        print("\n🤝 Топ Soft Skills:")
        for group_name, data in list(soft_skills_summary.items())[:8]:
            print(f"   {group_name} — ({data['count']})")

        print("\n" + "=" * 70)


# =============================================================================
# Блок для тестирования
# =============================================================================

if __name__ == "__main__":
    """Тестирование модуля расширенной аналитики."""

    import logging
    logging.getLogger("src").setLevel(logging.INFO)

    print("=" * 60)
    print("Тестирование AdvancedAnalytics")
    print("=" * 60)

    try:
        # Загружаем обработанные данные
        csv_path = settings.processed_data_dir / "vacancies_processed.csv"

        if csv_path.exists():
            print(f"\n📥 Загрузка данных из {csv_path}")
            df = pd.read_csv(csv_path)

            # Создаём аналитический модуль
            analytics = AdvancedAnalytics(df)

            # Выводим сводку
            analytics.print_advanced_summary()

            # Генерируем детальный отчёт
            print("\n📄 Генерация детального Excel-отчёта...")
            report_path = analytics.generate_detailed_excel_report()
            print(f"✅ Отчёт сохранён: {report_path}")

        else:
            print(f"⚠️  CSV файл не найден: {csv_path}")
            print("Сначала запустите processor.py")

    except KeyboardInterrupt:
        print("\n\n⚠️  Тестирование прервано пользователем")
    except Exception as e:
        print(f"\n❌ Ошибка: {type(e).__name__} - {e}")
        logger.exception("Детальная информация об ошибке")

    print("\n" + "=" * 60)
    print("Тестирование завершено!")
    print("=" * 60)
