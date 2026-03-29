"""
Модуль для аналитики и формирования отчётов.

Содержит класс VacancyAnalyzer, который:
- Анализирует обработанные данные из DataFrame
- Считает частоту упоминаний навыков
- Формирует отчёты в Excel (.xlsx)
- Генерирует консольные сводки

Важно: Для работы с Excel требуется openpyxl.
Отчёты сохраняются в data/reports/.
"""

from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
    Border,
    Side,
    Color,
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.config import settings, config_loader
from src.utils import get_logger, ensure_dir, format_number, safe_divide

# Инициализируем логгер для модуля
logger = get_logger(__name__)


class VacancyAnalyzer:
    """
    Анализатор вакансий для формирования отчётов.

    Реализует:
    - Подсчёт частоты навыков (Hard/Soft/Tools)
    - Анализ зарплат по различным группам
    - Формирование Excel-отчётов с графиками
    - Консольные сводки

    Атрибуты:
        df (pd.DataFrame): DataFrame с данными вакансий
        reports_dir (Path): Директория для сохранения отчётов

    Пример использования:
        >>> analyzer = VacancyAnalyzer(df)
        >>> analyzer.generate_excel_report()
        >>> analyzer.print_summary()
    """

    def __init__(
        self,
        df: pd.DataFrame,
        reports_dir: Optional[Path] = None
    ) -> None:
        """
        Инициализация анализатора.

        Args:
            df: DataFrame с обработанными вакансиями
            reports_dir: Директория для отчётов. Если None, берётся из конфига
        """
        self.df = df
        self.reports_dir = reports_dir or settings.reports_dir
        ensure_dir(self.reports_dir)

        # Кэш для статистики (чтобы не пересчитывать)
        self._skills_stats: Optional[Dict[str, Any]] = None

        logger.info(
            f"VacancyAnalyzer инициализирован. "
            f"Вакансий для анализа: {len(df)}"
        )

    def _count_skills(self, column: pd.Series) -> Dict[str, int]:
        """
        Подсчёт частоты навыков в колонке.

        Args:
            column: Серия с навыками (строки через запятую)

        Returns:
            Словарь {навык: частота}
        """
        skill_counts: Dict[str, int] = {}

        for skills_str in column.dropna():
            if isinstance(skills_str, str):
                # Разбиваем строку на навыки
                skills = [s.strip() for s in skills_str.split(",")]
                for skill in skills:
                    if skill:  # Пропускаем пустые
                        skill_counts[skill] = skill_counts.get(skill, 0) + 1

        return skill_counts

    def compute_skills_statistics(self) -> Dict[str, Any]:
        """
        Вычисление полной статистики по навыкам.

        Returns:
            Словарь со статистикой по всем категориям навыков

        Note:
            Результат кэшируется для повторного использования.
        """
        if self._skills_stats is not None and not self.df.empty:
            return self._skills_stats

        if self.df.empty:
            return {}

        # Считаем по каждой категории
        hard_skills_counts = self._count_skills(self.df["hard_skills"])
        soft_skills_counts = self._count_skills(self.df["soft_skills"])
        tools_counts = self._count_skills(self.df["tools"])

        # Сортируем по частоте
        top_n = config_loader.reporting.get("top_n_skills", 30)

        self._skills_stats = {
            "hard_skills": dict(sorted(
                hard_skills_counts.items(),
                key=lambda x: x[1],
                reverse=True
            )),
            "soft_skills": dict(sorted(
                soft_skills_counts.items(),
                key=lambda x: x[1],
                reverse=True
            )),
            "tools": dict(sorted(
                tools_counts.items(),
                key=lambda x: x[1],
                reverse=True
            )),
            "top_hard_skills": dict(sorted(
                hard_skills_counts.items(),
                key=lambda x: x[1],
                reverse=True
            )[:top_n]),
            "top_soft_skills": dict(sorted(
                soft_skills_counts.items(),
                key=lambda x: x[1],
                reverse=True
            )[:top_n]),
            "top_tools": dict(sorted(
                tools_counts.items(),
                key=lambda x: x[1],
                reverse=True
            )[:top_n]),
            "total_vacancies": len(self.df),
            "avg_skills_per_vacancy": self.df["skill_count"].mean(),
            "avg_hard_skills": self.df["hard_skill_count"].mean(),
            "avg_soft_skills": self.df["soft_skill_count"].mean(),
            "avg_tools": self.df["tools_count"].mean(),
        }

        return self._skills_stats

    def compute_salary_statistics(self) -> Dict[str, Any]:
        """
        Вычисление статистики по зарплатам.

        Returns:
            Словарь со статистикой зарплат по группам
        """
        if self.df.empty:
            return {}

        # Фильтруем вакансии с указанной зарплатой
        df_salary = self.df[self.df["salary_from"].notna()].copy()

        if df_salary.empty:
            return {"message": "Нет данных о зарплатах"}

        stats = {
            "total_with_salary": len(df_salary),
            "avg_salary_from": df_salary["salary_from"].mean(),
            "median_salary_from": df_salary["salary_from"].median(),
            "min_salary_from": df_salary["salary_from"].min(),
            "max_salary_from": df_salary["salary_from"].max(),
            "avg_salary_to": df_salary["salary_to"].mean() if "salary_to" in df_salary else None,
        }

        # Группировка по опыту работы
        if "experience" in self.df.columns:
            exp_groups = df_salary.groupby("experience")["salary_from"].agg(
                ["count", "mean", "median", "min", "max"]
            ).sort_values("count", ascending=False)

            stats["by_experience"] = exp_groups.to_dict("index")

        # Группировка по региону
        if "area" in self.df.columns:
            area_groups = df_salary.groupby("area")["salary_from"].agg(
                ["count", "mean", "median"]
            ).sort_values("count", ascending=False).head(20)

            stats["by_area"] = area_groups.to_dict("index")

        # Группировка по типу занятости
        if "employment" in self.df.columns:
            emp_groups = df_salary.groupby("employment")["salary_from"].agg(
                ["count", "mean", "median"]
            ).sort_values("count", ascending=False)

            stats["by_employment"] = emp_groups.to_dict("index")

        return stats

    def compute_experience_distribution(self) -> Dict[str, int]:
        """
        Распределение вакансий по требуемому опыту.

        Returns:
            Словарь {опыт: количество}
        """
        if self.df.empty or "experience" not in self.df.columns:
            return {}

        return self.df["experience"].value_counts().to_dict()

    def compute_area_distribution(self, top_n: int = 20) -> Dict[str, int]:
        """
        Распределение вакансий по регионам.

        Args:
            top_n: Количество топ регионов

        Returns:
            Словарь {регион: количество}
        """
        if self.df.empty or "area" not in self.df.columns:
            return {}

        return self.df["area"].value_counts().head(top_n).to_dict()

    def generate_excel_report(
        self,
        filename: Optional[str] = None
    ) -> Path:
        """
        Генерация полного Excel-отчёта.

        Args:
            filename: Имя файла. Если None, генерируется автоматически

        Returns:
            Путь к сохранённому файлу

        Note:
            Отчёт включает листы:
            - Summary (сводка)
            - Hard Skills
            - Soft Skills
            - Tools
            - Salaries
            - Experience
            - Areas
        """
        if self.df.empty:
            logger.warning("Нет данных для формирования отчёта")
            raise ValueError("DataFrame пуст")

        # Генерируем имя файла с timestamp
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"hh_analytics_report_{timestamp}.xlsx"

        filepath = self.reports_dir / filename

        logger.info(f"Генерация Excel-отчёта: {filepath}")

        # Создаём workbook
        wb = Workbook()

        # Удаляем стандартный лист
        wb.remove(wb.active)

        # Вычисляем статистику
        skills_stats = self.compute_skills_statistics()
        salary_stats = self.compute_salary_statistics()

        # Создаём листы
        self._create_summary_sheet(wb, skills_stats, salary_stats)
        self._create_skills_sheet(wb, "Hard Skills", skills_stats.get("hard_skills", {}))
        self._create_skills_sheet(wb, "Soft Skills", skills_stats.get("soft_skills", {}))
        self._create_skills_sheet(wb, "Tools", skills_stats.get("tools", {}))
        self._create_salary_sheet(wb, salary_stats)
        self._create_experience_sheet(wb)
        self._create_areas_sheet(wb)

        # Сохраняем
        wb.save(filepath)
        logger.info(f"Отчёт сохранён: {filepath}")

        return filepath

    def _create_summary_sheet(
        self,
        wb: Workbook,
        skills_stats: Dict[str, Any],
        salary_stats: Dict[str, Any]
    ) -> None:
        """Создание листа с общей сводкой."""
        ws = wb.create_sheet("Summary")

        # Стили
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", fill_type="solid")
        number_format = "#,##0"

        # Заголовок
        ws["A1"] = "HH.ru Analytics — Сводный отчёт"
        ws["A1"].font = Font(bold=True, size=18)
        ws.merge_cells("A1:B1")

        # Дата формирования
        ws["A3"] = "Дата формирования:"
        ws["B3"] = datetime.now().strftime("%d.%m.%Y %H:%M")

        # Общая статистика
        ws["A5"] = "Общая статистика"
        ws["A5"].font = header_font
        ws["A5"].fill = header_fill

        summary_data = [
            ["Всего вакансий", skills_stats.get("total_vacancies", 0)],
            ["Среднее навыков на вакансию", f"{skills_stats.get('avg_skills_per_vacancy', 0):.2f}"],
            ["Среднее Hard Skills", f"{skills_stats.get('avg_hard_skills', 0):.2f}"],
            ["Среднее Soft Skills", f"{skills_stats.get('avg_soft_skills', 0):.2f}"],
            ["Среднее Tools", f"{skills_stats.get('avg_tools', 0):.2f}"],
        ]

        if salary_stats:
            summary_data.extend([
                ["Вакансий с зарплатой", salary_stats.get("total_with_salary", 0)],
                ["Средняя зарплата (от)", f"{salary_stats.get('avg_salary_from', 0):.0f} RUB"],
                ["Медианная зарплата", f"{salary_stats.get('median_salary_from', 0):.0f} RUB"],
            ])

        # Записываем данные
        for i, (label, value) in enumerate(summary_data, start=6):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value

        # Топ-10 навыков
        ws["D5"] = "Топ-10 Hard Skills"
        ws["D5"].font = header_font
        ws["D5"].fill = header_fill

        top_hard = list(skills_stats.get("top_hard_skills", {}).items())[:10]
        ws[f"D6"] = "Навык"
        ws[f"E6"] = "Частота"
        ws["D6"].font = Font(bold=True)
        ws["E6"].font = Font(bold=True)

        for i, (skill, count) in enumerate(top_hard, start=7):
            ws[f"D{i}"] = skill
            ws[f"E{i}"] = count

        # Добавляем таблицу
        if top_hard:
            tab = Table(
                displayName="TopHardSkills",
                ref=f"D6:E{6 + len(top_hard)}"
            )
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            ws.add_table(tab)

    def _create_skills_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        skills_data: Dict[str, int]
    ) -> None:
        """Создание листа с навыками."""
        ws = wb.create_sheet(sheet_name)

        # Заголовок
        ws["A1"] = f"{sheet_name} — Частота упоминаний"
        ws["A1"].font = Font(bold=True, size=16)

        # Данные
        ws["A3"] = "Навык"
        ws["B3"] = "Частота"
        ws["C3"] = "% от вакансий"
        ws["A3"].font = Font(bold=True)
        ws["B3"].font = Font(bold=True)
        ws["C3"].font = Font(bold=True)

        total = self.df["skill_count"].sum() or 1

        for i, (skill, count) in enumerate(skills_data.items(), start=4):
            ws[f"A{i}"] = skill
            ws[f"B{i}"] = count
            ws[f"C{i}"] = f"{safe_divide(count * 100, total):.2f}%"

        # Добавляем таблицу
        if skills_data:
            tab = Table(
                displayName=f"{sheet_name.replace(' ', '')}Table",
                ref=f"A3:C{3 + len(skills_data)}"
            )
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showRowStripes=True
            )
            ws.add_table(tab)

        # Добавляем диаграмму (топ-15)
        top_15 = list(skills_data.items())[:15]
        if len(top_15) >= 3:
            chart = BarChart()
            chart.title = f"Топ-15 {sheet_name}"
            chart.style = 10
            chart.y_axis.title = "Частота"
            chart.x_axis.title = "Навык"

            # Данные для диаграммы
            data = Reference(
                worksheet=ws,
                min_col=2,
                min_row=3,
                max_row=3 + len(top_15),
                max_col=2
            )
            categories = Reference(
                worksheet=ws,
                min_col=1,
                min_row=4,
                max_row=3 + len(top_15)
            )

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.shape = 4
            ws.add_chart(chart, "E3")

    def _create_salary_sheet(
        self,
        wb: Workbook,
        salary_stats: Dict[str, Any]
    ) -> None:
        """Создание листа с зарплатами."""
        ws = wb.create_sheet("Salaries")

        # Заголовок
        ws["A1"] = "Анализ зарплат"
        ws["A1"].font = Font(bold=True, size=16)

        # Общая статистика
        ws["A3"] = "Общая статистика"
        ws["A3"].font = Font(bold=True, size=14)

        if salary_stats:
            summary = [
                ["Вакансий с зарплатой", salary_stats.get("total_with_salary", 0)],
                ["Средняя зарплата (от)", f"{salary_stats.get('avg_salary_from', 0):.0f} RUB"],
                ["Медианная зарплата", f"{salary_stats.get('median_salary_from', 0):.0f} RUB"],
                ["Минимальная", f"{salary_stats.get('min_salary_from', 0):.0f} RUB"],
                ["Максимальная", f"{salary_stats.get('max_salary_from', 0):.0f} RUB"],
            ]

            for i, (label, value) in enumerate(summary, start=4):
                ws[f"A{i}"] = label
                ws[f"B{i}"] = value

        # По опыту работы
        if salary_stats and "by_experience" in salary_stats:
            row = len(summary) + 6
            ws[f"A{row}"] = "По опыту работы"
            ws[f"A{row}"].font = Font(bold=True, size=14)
            row += 1

            ws[f"A{row}"] = "Опыт"
            ws[f"B{row}"] = "Вакансий"
            ws[f"C{row}"] = "Средняя"
            ws[f"D{row}"] = "Медиана"
            ws[f"A{row}"].font = Font(bold=True)

            row += 1
            for exp, data in salary_stats["by_experience"].items():
                ws[f"A{row}"] = exp
                ws[f"B{row}"] = int(data["count"])
                ws[f"C{row}"] = f"{data['mean']:.0f}"
                ws[f"D{row}"] = f"{data['median']:.0f}"
                row += 1

    def _create_experience_sheet(self, wb: Workbook) -> None:
        """Создание листа с распределением по опыту."""
        ws = wb.create_sheet("Experience")

        ws["A1"] = "Распределение по требуемому опыту"
        ws["A1"].font = Font(bold=True, size=16)

        distribution = self.compute_experience_distribution()

        ws["A3"] = "Опыт работы"
        ws["B3"] = "Количество вакансий"
        ws["A3"].font = Font(bold=True)
        ws["B3"].font = Font(bold=True)

        for i, (exp, count) in enumerate(distribution.items(), start=4):
            ws[f"A{i}"] = exp
            ws[f"B{i}"] = count

        # Диаграмма
        if len(distribution) >= 2:
            chart = PieChart()
            chart.title = "Распределение по опыту"
            chart.style = 10

            data = Reference(worksheet=ws, min_col=2, min_row=3, max_row=3 + len(distribution))
            categories = Reference(worksheet=ws, min_col=1, min_row=4, max_row=3 + len(distribution))

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            ws.add_chart(chart, "D3")

    def _create_areas_sheet(self, wb: Workbook) -> None:
        """Создание листа с регионами."""
        ws = wb.create_sheet("Areas")

        ws["A1"] = "Распределение по регионам"
        ws["A1"].font = Font(bold=True, size=16)

        distribution = self.compute_area_distribution()

        ws["A3"] = "Регион"
        ws["B3"] = "Количество вакансий"
        ws["A3"].font = Font(bold=True)
        ws["B3"].font = Font(bold=True)

        for i, (area, count) in enumerate(distribution.items(), start=4):
            ws[f"A{i}"] = area
            ws[f"B{i}"] = count

        # Таблица
        if distribution:
            tab = Table(
                displayName="AreasTable",
                ref=f"A3:B{3 + len(distribution)}"
            )
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showRowStripes=True
            )
            ws.add_table(tab)

    def print_summary(self) -> None:
        """
        Вывод сводной статистики в консоль.

        Note:
            Форматирует вывод для удобного чтения в терминале.
        """
        if self.df.empty:
            print("❌ Нет данных для анализа")
            return

        print("\n" + "=" * 60)
        print("📊 HH.ru Analytics — Сводка")
        print("=" * 60)

        skills_stats = self.compute_skills_statistics()
        salary_stats = self.compute_salary_statistics()

        # Общая статистика
        print(f"\n📈 Всего вакансий: {format_number(skills_stats.get('total_vacancies', 0))}")
        print(f"📋 Среднее навыков на вакансию: {skills_stats.get('avg_skills_per_vacancy', 0):.2f}")

        # Топ Hard Skills
        print("\n🛠 Топ-10 Hard Skills:")
        for i, (skill, count) in enumerate(
            list(skills_stats.get("top_hard_skills", {}).items())[:10], 1
        ):
            print(f"  {i:2}. {skill}: {format_number(count)}")

        # Топ Tools
        print("\n🔧 Топ-10 Tools:")
        for i, (skill, count) in enumerate(
            list(skills_stats.get("top_tools", {}).items())[:10], 1
        ):
            print(f"  {i:2}. {skill}: {format_number(count)}")

        # Зарплаты
        if salary_stats and "avg_salary_from" in salary_stats:
            print("\n💰 Зарплаты (RUB):")
            print(f"  Средняя: {salary_stats['avg_salary_from']:,.0f}")
            print(f"  Медиана: {salary_stats['median_salary_from']:,.0f}")

        print("\n" + "=" * 60)

    def export_to_csv(
        self,
        filename: Optional[str] = None
    ) -> Path:
        """
        Экспорт статистики в CSV.

        Args:
            filename: Имя файла. Если None, генерируется автоматически

        Returns:
            Путь к сохранённому файлу
        """
        if self.df.empty:
            raise ValueError("DataFrame пуст")

        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"hh_skills_stats_{timestamp}.csv"

        filepath = self.reports_dir / filename

        skills_stats = self.compute_skills_statistics()

        # Создаём DataFrame со статистикой
        rows = []
        for category, skills in [
            ("hard_skills", skills_stats.get("hard_skills", {})),
            ("soft_skills", skills_stats.get("soft_skills", {})),
            ("tools", skills_stats.get("tools", {}))
        ]:
            for skill, count in skills.items():
                rows.append({
                    "category": category,
                    "skill": skill,
                    "count": count
                })

        stats_df = pd.DataFrame(rows)
        stats_df.to_csv(filepath, index=False, encoding="utf-8-sig")

        logger.info(f"CSV сохранён: {filepath}")
        return filepath


# =============================================================================
# Блок для тестирования модуля
# =============================================================================

if __name__ == "__main__":
    """
    Пример использования VacancyAnalyzer для тестирования.

    Перед запуском убедитесь:
    1. Есть обработанный CSV файл в data/processed/
    2. Установлен openpyxl

    Запуск: python -m src.analyzer
    """

    import logging
    logging.getLogger("src").setLevel(logging.INFO)

    print("=" * 60)
    print("Тестирование VacancyAnalyzer")
    print("=" * 60)

    try:
        # Загружаем обработанные данные
        csv_path = settings.processed_data_dir / "vacancies_processed.csv"

        if csv_path.exists():
            print(f"\n📥 Загрузка данных из {csv_path}")
            df = pd.read_csv(csv_path)

            # Создаём анализатор
            analyzer = VacancyAnalyzer(df)

            # Выводим сводку
            analyzer.print_summary()

            # Генерируем Excel-отчёт
            print("\n📄 Генерация Excel-отчёта...")
            report_path = analyzer.generate_excel_report()
            print(f"✅ Отчёт сохранён: {report_path}")

            # Экспорт в CSV
            csv_stats_path = analyzer.export_to_csv()
            print(f"✅ CSV сохранён: {csv_stats_path}")

        else:
            print(f"⚠️  CSV файл не найден: {csv_path}")
            print("Сначала запустите processor.py")

    except KeyboardInterrupt:
        print("\n\n⚠️  Тестирование прервано пользователем")
    except Exception as e:
        print(f"\n❌ Ошибка: {type(e).__name__} - {e}")
        logger.exception("Детальная информация об ошибке")

    print("\n" + "=" * 60)
    print("Тестирование завершено!")
    print("=" * 60)
